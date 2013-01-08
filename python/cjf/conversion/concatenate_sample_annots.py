#! /usr/bin/env python
"""
utility function to concatenate all of the non-standard annotation fields of the samples in a .xls
or .xlsx tag file and creates a new .xlsx file with the concatenated fields in SAMPLE_CHARACTERISTICS

annotation fields are formated as follows:
column1:column1Val|column2:column2Val

USAGE:
concatenate_sample_annotations input_file output_file

VARIABLE DEFINITIONS:
input_file: the file to be operated on
output_file: the file to be created
"""
if __name__ == "main":
    def _is_number(s):
        """ utility function to check if the input string can be represented a number"""
        try:
            float(s)
            return True
        except ValueError:
            return False
    
    
    #imports
    import sys
    from openpyxl.writer.excel import ExcelWriter
    from openpyxl.reader.excel import load_workbook
    from openpyxl.cell import get_column_letter
    
    #parse the input arguments
    try:
        if len(sys.argv) == 2:
            wb1 = load_workbook(filename = sys.argv[1])
            ws1 = wb1.worksheets[0]
            output_path = sys.argv[1]
        elif len(sys.argv) == 3:
            wb1 = load_workbook(filename = sys.argv[1])
            ws1 = wb1.worksheets[0]
            output_path = sys.argv[2]
        else:
            raise NameError('Improper inputs')
    except IndexError:
        print (__doc__)
        sys.exit(0)
    except IOError:
        print (__doc__)
        sys.exit(0)
    
    
    #set up the notebook to write to
    ew = ExcelWriter(workbook = wb1)
    num_rows = len(ws1.rows) + 1
    num_cols = len(ws1.columns) + 1
    #find the column that has the SAMPLE_CHARACTERISTICS heading
    headers = [ws1.cell('%s1' %get_column_letter(col_ind) ).value for col_ind in range(1,num_cols)]
    SC_ind = headers.index('SAMPLE_CHARACTERISTICS')
    SC_letter = get_column_letter(SC_ind + 1)
    
    
    #concatenate the information from all columns that are not in the standard set of sixteen annotation
    #columns.  Restrict annotated data to those columns that do not appear in ignore_list and restrict
    #numeric fields to one place after the decimal point
    ignore_list = ['ARRAY', 'CHIP', 'ORG', 'CHARS', 'Present %', 'pct_present', 'all_signal']
    
    for row_ind in range(2,num_rows):
        annot_str = ''
        pipe_flag = 0
        for col_ind in range(17,num_cols):
            col_header = str(ws1.cell('%s1' %get_column_letter(col_ind) ).value)
            if col_header not in ignore_list:
                sample_val = str(ws1.cell('%s%s' %( get_column_letter(col_ind), str(row_ind)) ).value)
                if _is_number(sample_val):
                    if pipe_flag:
                        annot_str += '|'
                    annot_str += col_header.upper() + ':%.1f' % (float(sample_val))
                    pipe_flag = 1
                else:
                    if pipe_flag:
                        annot_str += '|'
                    annot_str += col_header.upper() + ':' + sample_val
                    pipe_flag = 1
        ws1.cell('%s%s' % (SC_letter,str(row_ind)) ).value = annot_str
    
    
    #write the new workbook to file
    ew.save(filename = output_path)
    print 'data written to %s' %(output_path)
