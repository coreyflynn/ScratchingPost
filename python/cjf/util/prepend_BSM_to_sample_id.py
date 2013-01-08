#! /usr/bin/env python
"""
utility funciton to prepend the "BSM_" prefix to all of the samples in the first column of a .xls
or .xlsx tag file and creates a new .xlsx file with the updated sample names

USAGE:
prepend_BSM_to_sample_id input_file output_file

VARIABLE DEFINITIONS:
input_file: the file to be operated on
output_file: the file to be created
"""

if __name__ == "main":
    #imports
    import sys
    
    print "importing openpyxl..."
    from openpyxl.writer.excel import ExcelWriter
    from openpyxl.reader.excel import load_workbook
    
    #parse the input arguments
    try:
        print "parsing %s..." %(sys.argv[1])
        wb1 = load_workbook(filename = sys.argv[1])
        ws1 = wb1.worksheets[0]
        output_path = sys.argv[2]
    except IndexError:
        print (__doc__)
        sys.exit(0)
    except IOError:
        print (__doc__)
        sys.exit(0)
    
    #set up the notebook to write to
    ew = ExcelWriter(workbook = wb1)
    num_rows = len(ws1.rows)
    
    #update the naming in the first column of samples
    print "updating sample names..."
    for row_ind in range(2,num_rows+1):
        if "BSM_" not in ws1.cell('A%s'% row_ind).value:
            ws1.cell('A%s'% row_ind).value = 'BSM_' + ws1.cell('A%s'% row_ind).value
    
    
    
    #write the new workbook to file
    print "writing %s..." %(sys.argv[2])
    ew.save(filename = sys.argv[2])
    